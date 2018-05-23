import openpyxl

catwb = openpyxl.load_workbook('Website_Categories.xlsx')
catws = catwb['Sheet1']

prodwb = openpyxl.load_workbook('BATHS MASTER_FINAL.xlsx')
prodws = prodwb['Sheet1']
templist = [[], []]

i=0
for row in prodws.iter_rows('V{}:W{}'.format(prodws.min_row, prodws.max_row)):
    for cell1 in row:
        if row[0].value.isdigit():
            templist[]
        # for row1 in catws.iter_rows('A{}:B{}'.format(catws.min_row, catws.max_row)):
        #     for cell2 in row1:
        #         if cell2.value == row1[0].value:


                    try:
                        print(46249)


                    except ValueError:
                        print("Value Error")

