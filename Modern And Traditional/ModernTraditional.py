import pypyodbc as pyodbc
import pprint
from collections import Counter
from openpyxl import Workbook

db_host = 'DESIGNER1\INTRANET'
db_name = 'intranet'
db_user = 'intranet'
db_password = 'bathcountry110'
connection_string = 'Driver={SQL Server};Server=' + db_host + ';Database=' + db_name + ';UID=' + db_user + ';PWD=' + db_password + ';'
db = pyodbc.connect(connection_string)
cursor = db.cursor()
pp = pprint.PrettyPrinter(indent=4)

cursor.execute('SELECT * from prod_parent_test')

columns = [column[0] for column in cursor.description]

# wb = Workbook()
# ws = wb.active


parentids = []

parentMTs = [[], []]

def checkEqual(iterator):
    iterator = iter(iterator)
    try:
        first = next(iterator)
    except StopIteration:
        return True
    return all(first == rest for rest in iterator)

def majority_element(lst):
    data = Counter(lst)
    return max(lst, key=data.get)

for row in cursor.fetchall():
    parentids.append(row[0])

currentChildren = []
# Iterate over all parent ids
iterator = 1
for id in parentids:

    # Select children with parentid = id
    cursor.execute('SELECT * from prod_child where displayproduct = 1 AND child_parentid = ' + str(id))
    childrows = []
    # Append all rows to childrows
    for row in cursor.fetchall():
        childrows.append(row)
        print(row)

    # tempChildren = []
    # for i in range(0, len(childrows)):
    #     tempChildren.append(childrows[i][24])
    # notequal = []
    # for child in tempChildren:
    #     if checkEqual(tempChildren) != True:
    #         ws['A' + str(iterator)] = id
    #         ws['B' + str(iterator)] = str(tempChildren)
    #         notequal.append(tempChildren)
    #         iterator += 1

        # if childrows[i].count(childrows[i][24]) != len(childrows):
        #     ws['A' + str(iterator)] = id
        #     ws['B' + str(iterator)] = childrows[i][0]

    # Iterate over child rows, add M or T to parentMTs if all modern_traditional children fields are equal or 0 if not equal
    for childrow in childrows:
        currentChildren.append(childrow[24])
        if childrows[iterator].count(childrow[24]) == len(childrows):
            parentMTs[0].append(id)
            parentMTs[1].append(majority_element(childrows)[24])
            break
        else:
            parentMTs[1].append(childrow[24])
            parentMTs[0].append(id)
            break

    # index = 0
    # for child in currentChildren:
    #     index += 1
    #     if currentChildren.count(child) != len(currentChildren):
    #         ws['A' + str(iterator)] = id
    #         ws['B' + str(iterator)] = childrow[0]
    #     if index == len(currentChildren):
    #         currentChildren = []

    # Iterate up until id 1000
    if id >= 1000:
        break
# wb.save('Non Identical Children.xlsx')
i = 0
for id in parentMTs[0]:
    if parentMTs[1][i] != None:
        cursor.execute('UPDATE prod_parent_test SET modern_traditional = \'' + parentMTs[1][i] + '\' WHERE parentid = '+str(id))
    i += 1
cursor.commit()

print(parentMTs)
print(len(parentMTs[0]))

cursor.close()
db.close()