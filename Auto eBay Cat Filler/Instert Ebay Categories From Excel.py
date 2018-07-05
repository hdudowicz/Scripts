import openpyxl
import pypyodbc as pyodbc
import re

db_host = 'DESIGNER1\INTRANET'
db_name = 'intranet'
db_user = 'intranet'
db_password = 'bathcountry110'
connection_string = 'Driver={SQL Server};Server=' + db_host + ';Database=' + db_name + ';UID=' + db_user + ';PWD=' + db_password + ';'
db = pyodbc.connect(connection_string)
cursor = db.cursor()

catwb = openpyxl.load_workbook('MASTER - CATEGORIES.xlsx')
catws = catwb['CORRELATION KEY']

prodwb = openpyxl.load_workbook('H:\Scripts\MASTER - SHOWER ENCLOSURES NEWW.xlsm')
prodws = prodwb['Sheet1']

tempwb = openpyxl.load_workbook('temp.xlsx')
tempws = tempwb['Sheet1']

prodcats = []

# Iterate over every SKU in product worksheet
for sku in prodws['C']:
    # Execute query subtracting the child part of the SKU
    cursor.execute("SELECT * FROM prod_parent WHERE prodcode = '" + re.sub(r'/\d.?', '', str(sku.value)) + "'")
    # cursor.execute("SELECT * FROM prod_parent WHERE prodcode = '" + str(sku.value) + "'")

    # Add relevant product category to prodcats array
    while True:
        row = cursor.fetchone()
        if not row:
            break
        prodcats.append(row['prod_parent_new_cat'])

i = 1

for prod in prodcats:
    print(catws['B' + str(prod)].value)
    tempws['A' + str(i)] = catws['B' + str(prod)].value
    tempws['B' + str(i)] = catws['D' + str(prod)].value

    tempws['D' + str(i)] = catws['H' + str(prod)].value
    tempws['E' + str(i)] = catws['J' + str(prod)].value
    i += 1
tempwb.save('Category List.xlsx')
